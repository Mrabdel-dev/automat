"""
thi file is for epessourage operation still in update and progress
"""
# import xlsxwriter
import xlsxwriter
from dbfread import DBF
from openpyxl import load_workbook

# load your pds file here
pdsFile = ''
pds = load_workbook('epesIn/pds-85_048_568.xlsx')
wpds = pds.sheetnames
# dbf file to get information about the boit
dbfFile = 'epesIn/21_011_079_BOITE_OPTIQUE_A.dbf'
boiteTable = DBF('epesIn/85_048_568_BOITE_OPTIQUE_A.dbf', load=True, encoding='iso-8859-1')
filedBoiteNam = boiteTable.field_names
boiteLen = len(boiteTable)
boiteCode = []
codeLocal = []
for j in range(0, boiteLen):
    boiteCode.append(boiteTable.records[j]['NOM'])
    codeLocal.append(boiteTable.records[j]['ID_PARENT'])
# create the epesourege file
epesBook = xlsxwriter.Workbook('epesExcel/SRO-85_048_568_568Epesourage.xlsx')
wr = epesBook.add_worksheet()
print(wpds)
boiteList = sorted(wpds)
print(boiteList)
header = epesBook.add_format({'bold': True, 'border': 1, 'bg_color': '#037d50'})
border = epesBook.add_format({"border": 1})
# ################# the part of coping values from pds to new file ######################
wr.write('A1', 'CODE_CABLE_ORIGINE', header)
wr.write('B1', 'NUMERO_TUBE_ORIGINE', header)
wr.write('C1', 'BAGUE_TUBE_ORIGINE ', header)
wr.write('D1', 'COULEUR_TUBE_ORIGINE', header)
wr.write('E1', 'NUMERO_FIBRE_ORIGINE', header)
wr.write('F1', 'BAGUE_FIBRE_ORIGINE ', header)
wr.write('G1', 'COULEUR_FIBRE_ORIGINE', header)
wr.write('H1', 'LOVAGE_FIBRE_ORIGINE', header)
wr.write('I1', 'CODE_SITE', header)
wr.write('J1', 'CODE_NIVEAU', header)
wr.write('K1', 'CODE_LOCALTECHNIQUE', header)
wr.write('L1', 'CODE_BOITE', header)
wr.write('M1', 'CODE_CASSETTE', header)
wr.write('N1', 'POSITION_CASSETTE', header)
wr.write('O1', 'CODE_CABLE_DESTINATION', header)
wr.write('P1', 'NUMERO_TUBE_DESTINATION', header)
wr.write('Q1', 'BAGUE_TUBE_DESTINATION ', header)
wr.write('R1', 'COULEUR_TUBE_DESTINATION', header)
wr.write('S1', 'NUMERO_FIBRE_DESTINATION', header)
wr.write('T1', 'BAGUE_FIBRE_DESTINATION ', header)
wr.write('U1', 'COULEUR_FIBRE_DESTINATION', header)
wr.write('V1', 'LOVAGE_FIBRE_DESTINATION', header)
wr.write('W1', 'ETAT', header)
b = 2


def integerFormat(x):
    test = str(x)
    if test.isdigit():
        f = 'CSE-' + test.zfill(2)
        return f

    elif test.startswith('FON'):
        test = 'FOND DE BOITE'
        return test
    else:
        return x


def getBagueByTube(tube: str):
    if tube.isdigit():
        tube = int(tube)
        if tube <= 12:
            return 1
        elif tube <= 24:
            return 2
        elif tube <= 36:
            return 3
        elif tube <= 48:
            return 4
        elif tube <= 60:
            return 5
        else:
            return 6
    else:
        return ''


code = ''
for s in boiteList:
    sheet = pds[s]
    MaxRow = sheet.max_row
    print(MaxRow)
    MaxCol = sheet.max_column
    for t in range(0, boiteLen):
        if s == boiteCode[t]:
            code = codeLocal[t]

    for i in range(2, MaxRow + 1):
        cable = sheet.cell(row=i, column=1).value
        wr.write('A' + str(b), cable, border)
        # tube1
        tube1 = sheet.cell(row=i, column=4).value
        wr.write('B' + str(b), tube1, border)
        # bugue
        x = getBagueByTube(str(tube1))
        if x is not None:
            wr.write('C' + str(b), x, border)
        else:
            wr.write('C' + str(b), '', border)

        wr.write('D' + str(b), '', border)
        # fibre1
        fibre1 = sheet.cell(row=i, column=5).value
        wr.write('E' + str(b), fibre1, border)
        wr.write('F' + str(b), '', border)
        wr.write('G' + str(b), '', border)
        wr.write('H' + str(b), '', border)
        wr.write('I' + str(b), '', border)
        wr.write('J' + str(b), '', border)
        # CODElOCAL
        wr.write('K' + str(b), code, border)
        # boite
        wr.write('L' + str(b), s, border)
        # cassete
        cassete = sheet.cell(row=i, column=6).value
        wr.write('M' + str(b), integerFormat(cassete), border)
        # position
        position = sheet.cell(row=i, column=3).value
        wr.write('N' + str(b), position, border)
        # cable dist
        cableDist = sheet.cell(row=i, column=13).value
        wr.write('O' + str(b), cableDist, border)
        # tube2
        tube2 = sheet.cell(row=i, column=9).value
        wr.write('P' + str(b), tube2, border)
        # bague 2
        x = getBagueByTube(str(tube2))
        if x is not None:
            wr.write('Q' + str(b), x, border)
        else:
            wr.write('Q' + str(b), '', border)

        wr.write('R' + str(b), '', border)
        # fibre 2
        fibre2 = sheet.cell(row=i, column=8).value
        wr.write('S' + str(b), fibre2, border)
        wr.write('T' + str(b), '', border)
        wr.write('U' + str(b), '', border)
        wr.write('V' + str(b), '', border)
        # ETAT
        type = sheet.cell(row=i, column=7).value
        if type == 'EN ATTENTE' or type == 'PASSAGE':
            type = 'EN PASSAGE'
        elif type == 'LIBRE' or type == 'A STOCKER' or type == 'STOCKER':
            type = 'STOCKEE'
        elif type == 'A EPISSURER' or type == 'EPISSURER' or type == 'A EPISSUREE':
            type = 'EPISSUREE'
        wr.write('W' + str(b), type, border)
        b += 1

epesBook.close()
