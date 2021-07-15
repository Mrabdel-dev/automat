""
import xlsxwriter
from PIL import Image
from os import walk

# the folder source that you want take resize inside it
from openpyxl import load_workbook

monRepertoire = 'C:/Users/etudes20/Desktop/FCI 075/'
# the folder out that u want to put new images in it
sortRepetoire = 'C:/Users/etudes20/Desktop/FCI 075/fCI-075.xslx'
workbook1 = xlsxwriter.Workbook('fCI-075.xlsx')
w = workbook1.add_worksheet('FCI-075')
bold = workbook1.add_format({'bold': True, "border": 1})
bold1 = workbook1.add_format({'bold': True})
border = workbook1.add_format({"border": 1})
w.write(0, 0, 'N° ', bold)
w.write(0, 1, 'FCI', bold)
listP = set()
listeFichiers = []
for (repertoire, sousRepertoires, fichiers) in walk(monRepertoire):
    listeFichiers.extend(fichiers)

i = 0
t = 1
print(len(listeFichiers))
while i < len(listeFichiers):
    workbook = load_workbook(monRepertoire + listeFichiers[i])
    BookC3 = workbook['Commandes Fermes']
    maxRow = BookC3.max_row
    fci = str(BookC3.cell(row=4, column=3).value)
    print(fci)
    for k in range(15, maxRow):
        a = str(BookC3.cell(row=k, column=3).value)
        b = str(BookC3.cell(row=k, column=5).value)
        if a is not None :
            listP.add(a)
        if b is not None:
            listP.add(b)
    for p in listP:
        w.write(t, 0, p, border)
        w.write(t, 1, fci, border)
        t += 1
    listP = set()
    i += 1
workbook1.close()