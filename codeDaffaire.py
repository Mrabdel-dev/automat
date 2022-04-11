import xlsxwriter
from PIL import Image
from os import walk

# the folder source that you want take resize inside it
from openpyxl import load_workbook
num='854'
monRepertoire = 'C:/Users/etudes20/Desktop/desktop 1/c3a/code daffaires/'
# the folder out that u want to put new images in it
sortRepetoire = f'C:/Users/etudes20/Desktop/desktop 1/c3a/code d-affaire {num}.xlsx'
workbook1 = xlsxwriter.Workbook(sortRepetoire)
w = workbook1.add_worksheet('code')
bold = workbook1.add_format({'bold': True, "border": 1})
bold1 = workbook1.add_format({'bold': True})
border = workbook1.add_format({"border": 1})
w.write(0, 0, 'NOM', bold)
w.write(0, 1, 'code', bold)
listP = set()
listeFichiers = []
for (repertoire, sousRepertoires, fichiers) in walk(monRepertoire):
    listeFichiers.extend(fichiers)

i = 0
t = 1
print(len(listeFichiers))
while i < len(listeFichiers):
    workbook = load_workbook(monRepertoire + listeFichiers[i])
    BookC3 = workbook.active
    maxRow = BookC3.max_row
    ensee = str(BookC3.cell(row=1, column=9).value)+"/"
    print(ensee)
    for k in range(4, maxRow):
        a = str(BookC3.cell(row=k, column=1).value)

        if a is not None and a.startswith("B"):
            a = a[2:]
            print(a)
            listP.add(ensee+a)
    code = listeFichiers[i][0:4]+"-"+listeFichiers[i][4:-5]
    for p in listP:
        w.write(t, 0, p, border)
        w.write(t, 1,code, border)
        t += 1
    listP = set()
    i += 1
workbook1.close()
