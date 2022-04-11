from openpyxl import load_workbook

# load the THE ROUTE FILE
workbook = load_workbook('netgeo/Rootage-SRO-21_017_101.xlsx')
rout = workbook.active
# load the ROP FILE
workbook1 = load_workbook('netgeo/RO_SRO-21-014-101_V2.xlsx')
rop = workbook1["Sheet1"]

maxRowRoute = rout.max_row + 1
maxColRoute = rout.max_column + 1
result = []

for r in range(2, maxRowRoute):
    for c in range(8, maxColRoute):
        val = str(rout.cell(r, c).value)
        val2 = str(rop.cell(r, c).value)
        if val != val2 and not val.startswith("CSE") and not val.startswith("F"):
            print(val, val2, " error  in ligne : " + str(r) + "and col " + str(c))
            break
