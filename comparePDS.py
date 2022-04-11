from os import walk
import xlsxwriter
from openpyxl import load_workbook
workbook = xlsxwriter.Workbook(f'PDS/PDSg.xlsx')
border = workbook.add_format({"border": 1})
back = workbook.add_format({"bg_color": '#CD5C5C', "border": 1})
header = workbook.add_format({'bold': True, 'border': 1, 'bg_color': '#C4E5F7'})
cassette = workbook.add_format({"bg_color": '#A9A9A9', "border": 1})
cell_formatCapacity = workbook.add_format({"bg_color": '#E6E6FA', "border": 1})
cell_format1 = workbook.add_format({"bg_color": 'red', "border": 1})
cell_format2 = workbook.add_format({"bg_color": 'blue', "border": 1})
cell_format3 = workbook.add_format({"bg_color": '#00FF00', "border": 1})
cell_format4 = workbook.add_format({"bg_color": 'yellow', "border": 1})
cell_format5 = workbook.add_format({"bg_color": '#BF00FF', "border": 1})
cell_format6 = workbook.add_format({"bg_color": 'white', "border": 1})
cell_format7 = workbook.add_format({"bg_color": '#FFBF00', "border": 1})
cell_format8 = workbook.add_format({"bg_color": '#828282', "border": 1})
cell_format9 = workbook.add_format({"bg_color": '#816B56', "border": 1})
cell_format10 = workbook.add_format({"bg_color": '#333333', "border": 1})
cell_format11 = workbook.add_format({"bg_color": '#00FFBF', "border": 1})
cell_format12 = workbook.add_format({"bg_color": '#FFAAD4', "border": 1})
colorList = [cell_format1, cell_format2, cell_format3, cell_format4, cell_format5, cell_format6, cell_format7,
             cell_format8, cell_format9, cell_format10, cell_format11, cell_format12, border]
def stringCassette(x: str):
    if x is None:
        return colorList[12]
    if x.isdigit():
        j = 0
        if int(x) % 12 == 0:
            x = 12
        else:
            x = int(x) % 12

        for i in range(0, 13):
            if i == x:
                x = i
                j = 1
        if j == 1:
            return colorList[x - 1]
        else:
            return colorList[12]
    return colorList[12]
rep2 = 'C:/Users/etudes20/Desktop/mon plan/'
rep1 = 'C:/Users/etudes20/Desktop/plans_de_boites/'
listeFichiers1 = []
for (repertoire, sousRepertoires, fichiers) in walk(rep1):
    listeFichiers1.extend(fichiers)

listeFichiers2 = []
for (repertoire, sousRepertoires, fichiers) in walk(rep2):
    listeFichiers2.extend(fichiers)

i = 0
while i < len(listeFichiers1):
    workbook1 = load_workbook(rep1 + listeFichiers1[i])
    w1 = workbook1.active
    # w = workbook.add_worksheet(str(listeFichiers1[i])[0:-5])
    workbook2 = load_workbook(rep2 + listeFichiers1[i])
    w2 = workbook2.active
    maxrow = w1.max_row
    maxcol = w1.max_column
    for r in range(12,maxrow):
        cable1 = str(w1.cell(r, 2).value)
        tube1 = str(w1.cell(r, 3).value)
        fibre1 = str(w1.cell(r, 4).value)
        etat1 = str(w1.cell(r, 5).value)
        cassete1 = str(w1.cell(r, 6).value)
        position1 = str(w1.cell(r, 7).value)
        fib1D = str(w1.cell(r, 8).value)
        tube1d = str(w1.cell(r, 9).value)
        if tube1d is None or fib1D is None:
            tube1d =""
            fib1D =""
        ##########
        cable2 = str(w2.cell(r, 2).value)
        tube2 = str(w2.cell(r, 3).value)
        fibre2 = str(w2.cell(r, 4).value)
        etat2 = str(w2.cell(r, 5).value)
        cassete2 = str(w2.cell(r, 6).value)
        position2 = str(w2.cell(r, 7).value)
        fib2D = str(w2.cell(r, 8).value)
        tube2d = str(w2.cell(r, 9).value)
        if tube2d is None or tube2d == "None" or  fib2D is None =="None":
            print("yes")
            tube2d = ""
            fib2D = ""
        if cable1 != cable2 :
            print("eroooooooooor cble ", listeFichiers1[i], " at line ", r," ",cable1," ",cable2)
        elif cassete1 != cassete2 or position1 != position2:
            print("eroooooooooor casstee ", listeFichiers1[i], " at line ", r, "casstte1 ", cassete1, " casste2 ",
                  cassete2)
        elif tube1 != tube2 or fibre1 != fibre2 :
            print("eroooooooooor tube or fibre ", listeFichiers1[i], " at line ", r," ",fibre1," ",fibre2)
        elif etat1 != etat2:
            print("eroooooooooor etat ", listeFichiers1[i], " at line ", r," ",etat1," ",etat2)
        elif fib1D != fib2D or tube1d != tube2d:
            print("eroooooooooor tube2 ", listeFichiers1[i], " at line ", r," ",tube1d," ",fib1D," ",tube2d," ",fib2D)
    i +=1
