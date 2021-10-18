import os
from openpyxl import load_workbook

rep = 'C:/Users/etudes20/Desktop/1_FOA/'

workbook = load_workbook('C:/Users/etudes20/Downloads/foa.xlsx')
sheet = workbook.active
maxrow = sheet.max_row

filename = []
for i in range(1, maxrow):
    x = str(sheet.cell(row=i, column=1).value)
    if x != "":
        filename.append(x)
print("#" * 18, len(filename), "#" * 8)
for k in filename:
    try:
        print(k, "was deleted")
        os.remove(rep + str(k) + ".xlsx")
    except:
        print(k, "not deleted")
