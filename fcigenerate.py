from dbfread import DBF
from openpyxl import load_workbook

workbook = load_workbook('FCI 21-011/fCI-011.xlsx')
sheet = workbook.active
maxrow = sheet.max_row
# ####################### load the both file boite and cable in DBF format ###################################
supportTable = DBF('FCI 21-011/21_011_SUPPORT_A.dbf', load=True, encoding='iso-8859-1')
pointTechTable = DBF('FCI 21-011/21_011_POINT_TECHNIQUE_A.dbf', load=True, encoding='iso-8859-1')
pointlen = len(pointTechTable)
supplen = len(supportTable)
# FROM THE TECHNIC POINT
pointNom = []
pointCode = []
pointStruc = []
pointPrp = []
pointChambre = []
for p in range(0, pointlen):
    x = pointTechTable.records[p]['TYPE_STRUC']
    pointStruc.append(x)
    pointNom.append(pointTechTable.records[p]['NOM'])
    pointCode.append(pointTechTable.records[p]['CODE'])
    pointPrp.append(pointTechTable.records[p]['PROPRIETAI'])
    if str(x).startswith('CH'):
        pointChambre.append(pointTechTable.records[p]['NOM'])

# FROM THE SUPPORT
suppAmount = []
suppAval = []
for s in range(0, supplen):
    suppAmount.append(supportTable.records[s]['AMONT'])
    suppAval.append(supportTable.records[s]['AVAL'])

nom = []
fci = []
print(maxrow)
for i in range(2, maxrow):
    nom.append(sheet.cell(row=i, column=1).value)
    fci.append(sheet.cell(row=i, column=2).value)


def checkfci(point):
    try:
        if point.startswith('CH'):
            point = point[4:9] + '/' + point[10:]
        index = nom.index(point)
        return True
    except ValueError:
        return False


def getFci(point):
    if point.startswith('CH'):
        point = point[4:9] + '/' + point[10:]
    index = nom.index(point)
    return fci[index]


def getamont(point):
    try:
        ind = suppAval.index(point)
        return suppAmount[ind]
    except:
        ind = suppAmount.index(point)
        return suppAval[ind]


def getAval(point):
    try:
        ind = suppAmount.index(point)
        return suppAval[ind]
    except ValueError:
        ind = suppAval.index(point)
        return suppAmount[ind]


def getCloseFci(point):
    x = getamont(point)

    if x.startswith('S'):
        prop = 'Ene'
    else:
        ind = pointNom.index(x)
        prop = pointPrp[ind]
    if prop.startswith('OR'):
        if checkfci(x):
            return x
    else:
        x = getAval(point)
        ind = pointNom.index(x)
        prop = pointPrp[ind]
        if prop.startswith('OR'):
            if checkfci(x):
                return x
            else:
                return getCloseFci(x)
        else:
            point = getamont(point)
            return getCloseFci(point)


k = maxrow + 1
print(pointChambre,pointChambre.index('CHA-21001-4875'))

for p in pointChambre:
    ind = pointNom.index(p)
    prop = pointPrp[ind]
    if checkfci(p) or prop.startswith('CON'):
        continue
    else:
        try:
            x = getCloseFci(p)
            fcic = getFci(x)
            fci.append(fcic)
            nom.append(p)
            print("for point-->", p, " give this fci ", x, " -->", fcic)
            sheet.cell(row=k, column=1).value = p
            sheet.cell(row=k, column=2).value = fcic
            k += 1
        except:
            print(p)

workbook.save('FCI 21-011/fCI-011-New.xlsx')
