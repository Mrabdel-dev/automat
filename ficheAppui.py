from os import walk

from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# hna fin kaynin les photos
rep = 'C:/Users/etudes20/Desktop/photos/'
# hna dossier des fichier // note ---> remove fichier c6
monRepertoire = r'C:\Users\etudes20\Desktop\filr/'
ok = 'C:/Users/etudes20/Desktop/ok/'
rec = 'C:/Users/etudes20/Desktop/rec/'
rem = 'C:/Users/etudes20/Desktop/rem/'
save=""
listeFichiers = []
for (repertoire, sousRepertoires, fichiers) in walk(monRepertoire):
    listeFichiers.extend(fichiers)
i = 0
ext = '.JPG'
while i < len(listeFichiers):
    try:
        workbook = load_workbook(monRepertoire + listeFichiers[i])
    except :
        print(listeFichiers[i])
        i+=1
        continue
    x = str(listeFichiers[i])[11:-5]
    print(x)
    BookC6 = workbook[x]
    typeAp = str(BookC6.cell(row=53, column=13).value)
    if typeAp.startswith("Rem"):
        save = rem
    elif typeAp.startswith("Rem"):
        save=rec
    else:
        save = ok
    try:
        try:
            img = Image(rep + x + "_1" + ext)
            img.height = 308
            img.width = 360
            BookC6.add_image(img, "A" + str(62))
            img2 = Image(rep + x + "_2" + ext)
            img2.height = 308
            img2.width = 360
            BookC6.add_image(img2, "M" + str(62))
            if typeAp.startswith("Rem"):
                img3 = Image(rep + x + "_3" + ext)
                img3.height = 308
                img3.width = 360
                BookC6.add_image(img4, "A" + str(62))
                img4 = Image(rep + x + "_4" + ext)
                img4.height = 308
                img4.width = 360
                BookC6.add_image(img4, "M" + str(62))

        except FileNotFoundError:
            print(listeFichiers[i])
            img = Image(rep + x[1:] + "_1" + ext)
            img.height = 308
            img.width = 360
            BookC6.add_image(img, "A" + str(62))
            img2 = Image(rep + x[1:] + "_2" + ext)
            img2.height = 308
            img2.width = 360
            BookC6.add_image(img2, "M" + str(62))
            if typeAp.startswith("Rem"):
                img3 = Image(rep + x[1:] + "_3" + ext)
                img3.height = 308
                img3.width = 360
                BookC6.add_image(img4, "A" + str(62))
                img4 = Image(rep + x[1:] + "_4" + ext)
                img4.height = 308
                img4.width = 360
                BookC6.add_image(img4, "M" + str(62))
    except :
        print(listeFichiers[i])

    workbook.save(save + listeFichiers[i])
    i += 1

