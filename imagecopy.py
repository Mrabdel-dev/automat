from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# load the C6 book
rep = 'C:/Users/etudes20/Desktop/Exemple pour abdellah/photos/'
workbook = load_workbook('C:/Users/etudes20/Desktop/Exemple pour abdellah/CAPFT_085877-085891.xlsx')
BookC6 = workbook['Photos']
maxraw = BookC6.max_row
ext = '.JPG'
val = 3
t = 2
for i in range(0, maxraw):
    name = str(BookC6.cell(row=val, column=1).value)
    try:
        img = Image(rep + name + ext)
        img.height = 400
        img.width = 320

        BookC6.add_image(img, "A" + str(t))
    except FileNotFoundError:
        print(name)
    name = str(BookC6.cell(row=val, column=2).value)
    try:
        img = Image(rep + name + ext)
        img.height = 400
        img.width = 320
        BookC6.add_image(img, "B" + str(t))
    except FileNotFoundError:
        print(name)
    val += 2
    t += 2
workbook.save('CAPFT_.xlsx')
