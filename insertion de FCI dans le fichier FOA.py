# -*- coding: utf-8 -*-
"""
Created on Thu Mar  4 17:49:03 2021

@author: BE24
"""
import openpyxl
from os import walk

monRepertoire = 'C:\\Users\\bureauetude03\\Desktop\\Ali\\nomage FCI\\releve de chambre\\'

listeFichiers = []
for (repertoire, sousRepertoires, fichiers) in walk(monRepertoire):
 listeFichiers.extend(fichiers)
 
i = 0
while i < len(listeFichiers):
    wb = openpyxl.load_workbook(monRepertoire + listeFichiers[i])
    ws = wb.active
    try:
        ws.unmerge_cells('E2:P2')
       
        ws["E2"] = 'F09892110321'
        
        ws.merge_cells('E2:P2')
        
        wb.save('C:\\Users\\bureauetude03\\Desktop\\Ali\\nomage FCI\\rlc\\' + listeFichiers[i] )
    except:

        ws["E2"] = 'F09892110321'
        
        ws.merge_cells('E2:P2')
        
        wb.save('C:\\Users\\bureauetude03\\Desktop\\Ali\\nomage FCI\\rlc\\' + listeFichiers[i] )   
    i +=1

