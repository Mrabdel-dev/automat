import datetime

import xlsxwriter
from openpyxl import load_workbook

# load the THE ROUTE FILE
file = "Rootage-SRO-85_072_794"
workbookd = load_workbook('routage/Rootage-SRO-85_072_794.xlsx')
rout = workbookd.active
maxRowRoute = rout.max_row + 1
maxColRoute = rout.max_column + 1

workbook = xlsxwriter.Workbook(f'netgeo/{file} version losange.xlsx')
sheet = xlsxwriter.worksheet.Worksheet
# ############### define the character and style of cell inside excel ################"
bold = workbook.add_format({'bold': True, "border": 1})
bold1 = workbook.add_format({'bold': True})
border = workbook.add_format({"border": 1})
back = workbook.add_format({"bg_color": '#CD5C5C', "border": 1})
header = workbook.add_format({'bold': True, 'border': 1, 'bg_color': '#C4E5F7'})
cassette = workbook.add_format({"bg_color": '#A9A9A9', "border": 1})
cell_formatCapacity = workbook.add_format({"bg_color": '#E6E6FA', "border": 1})
cell_format1 = workbook.add_format({"bg_color": 'red', "border": 1})
cell_format2 = workbook.add_format({"bg_color": 'blue', "border": 1})
cell_format3 = workbook.add_format({"bg_color": '#00FF00', "border": 1})
cell_format4 = workbook.add_format({"bg_color": 'yellow', "border": 1})
cell_format5 = workbook.add_format({"bg_color": '#BF00FF', "border": 1})
cell_format6 = workbook.add_format({"bg_color": 'white', "border": 1})
cell_format7 = workbook.add_format({"bg_color": '#FFBF00', "border": 1})
cell_format8 = workbook.add_format({"bg_color": '#828282', "border": 1})
cell_format9 = workbook.add_format({"bg_color": '#816B56', "border": 1})
cell_format10 = workbook.add_format({"bg_color": '#333333', "border": 1})
cell_format11 = workbook.add_format({"bg_color": '#00FFBF', "border": 1})
cell_format12 = workbook.add_format({"bg_color": '#FFAAD4', "border": 1})
colorList = [cell_format1, cell_format2, cell_format3, cell_format4, cell_format5, cell_format6, cell_format7,
             cell_format8, cell_format9, cell_format10, cell_format11, cell_format12, border]


def stringCassette(x: str):
    if x.isdigit():
        j = 0
        if int(x) % 12 == 0:
            x = 12
        else:
            x = int(x) % 12

        for i in range(0, 13):
            if i == x:
                x = i
                j = 1
        if j == 1:
            return colorList[x - 1]
        else:
            return colorList[12]
    return colorList[12]


w = workbook.add_worksheet("result")
line = 1
w.write('A' + str(line), 'Connecteur_Origine', header)
w.write('B' + str(line), 'Code_Destination', header)
w.write('C' + str(line), 'Plateau_Origine', header)
w.write('D' + str(line), 'Plateau', header)
w.write('E' + str(line), 'Connecteur', header)
w.write('F' + str(line), 'Tube_Extremite', header)
w.write('G' + str(line), 'Fibre_Extremite', header)
w.write('H' + str(line), 'long_carto', header)
w.write('I' + str(line), 'long_calc', header)
w.write('J' + str(line), 'N°reflecto', header)
w.write('K' + str(line), 'TCO', header)
w.write('L' + str(line), 'Position SRO', header)
w.write('M' + str(line), 'Distance mesuré', header)
w.write('N' + str(line), 'Commentaire', header)
line = 2
ct = 1
# w.set_column(0, 20, 44)
for r in range(2, maxRowRoute):
    teroir = str(rout.cell(r, 5).value)[-1]
    cass = str(rout.cell(r, 7).value)
    p = str(rout.cell(r, 4).value)
    refelct = str(rout.cell(r, 2).value)
    for c in range(12, maxColRoute):
        val = str(rout.cell(r, c).value)

        if val == "STOCKEE" or val.startswith("S"):
            boite = str(rout.cell(r, c - 1).value)

            fibre = str(rout.cell(r, c - 3).value)
            tube = str(rout.cell(r, c - 4).value)
            w.write('A' + str(line), 'CO' + str(ct).zfill(9), border)
            ct += 1
            w.write("B" + str(line), boite, border)
            w.write("C" + str(line), 'TDI-21-011-104-01-A2-0' + teroir, border)
            w.write("D" + str(line), p, border)
            w.write("E" + str(line), cass, border)
            w.write("F" + str(line), tube, stringCassette(tube))
            w.write("G" + str(line), fibre, stringCassette(fibre))
            w.write("H" + str(line), '', border)
            w.write("I" + str(line), '', border)
            w.write("J" + str(line), refelct, border)
            w.write("K" + str(line), '', border)
            w.write("L" + str(line), 'T' + teroir + ' C' + p + ' F' + cass, border)
            w.write("M" + str(line), '', border)
            w.write("N" + str(line), '', border)
            line += 1

workbook.close()
