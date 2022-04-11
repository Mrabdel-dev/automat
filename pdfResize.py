from openpyxl import load_workbook
from openpyxl.drawing.image import Image
fileC6 = r'C:\Users\etudes20\Desktop\C6-85-081-856-3(1).xlsx'
rep = r'C:\Users\etudes20\Desktop\Photos sro856\PHOTOS SRO 856/'

annexeC6 = load_workbook(fileC6)
export = annexeC6['Export 1']
maxRow = export.max_row
photo = annexeC6['Photos']

ensseCode = export.cell(row=3, column=7).value
print(ensseCode)
lin = 3
for i in range(9, maxRow + 1):
    val = export.cell(row=i, column=1).value
    if val is not None and len(val) < 8:
        photo.cell(row=lin, column=1).value = val + "_" + "1"
        photo.cell(row=lin, column=2).value =  val + "_" + "2"
        lin += 2

annexeC6.save(fileC6)
workbook = load_workbook(fileC6)
BookC6 = workbook['Photos']
maxraw = BookC6.max_row
ext = '.JPG'
val = 3
t = 2
for i in range(0, maxraw):
    name = str(BookC6.cell(row=val, column=1).value)
    try:
        img = Image(rep + name + ext)
        img.height = 400
        img.width = 320

        BookC6.add_image(img, "A" + str(t))
    except FileNotFoundError:
        print(name)
    name = str(BookC6.cell(row=val, column=2).value)
    try:
        img = Image(rep + name + ext)
        img.height = 400
        img.width = 320
        BookC6.add_image(img, "B" + str(t))
    except FileNotFoundError:
        print(name)
    val += 2
    t += 2
workbook.save(fileC6)