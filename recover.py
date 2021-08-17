# just try to learn well
from tkinter import *
from tkinter import filedialog

top = Tk()
top.title("Maneo File Generator")
top.minsize(800, 400)


def browsefunc():
    filename = filedialog.askopenfilename()
    pathlabel.config(text=filename)


browsebutton = Button(top, text="Browse", command=browsefunc)
browsebutton.pack()

pathlabel = Label(top)
pathlabel.pack()
top.mainloop()
# def duplicates(lst, item):
#     return [i for i, x in enumerate(lst) if x == item]
#
#
#
# listi = ['1','2','158','99','93']
# for i in enumerate(listi):
#     print(i)
# c = []
# for i in range(65, 77):
#     c.append(chr(i))
# c.append(chr(78))
# import openpyxl.worksheet.worksheet
# import xlsxwriter.worksheet

from openpyxl import load_workbook

# pdsBook = load_workbook('fileGenerated/PDS.xlsx')
# pdsSheets = pdsBook.sheetnames
# y = pdsBook[pdsSheets[0]]
# print(type(y))
# def sheets(x:openpyxl.worksheet.worksheet.Worksheet):
#     x.cell()
# print(c[15])
# import xlsxwriter
# rootBook = xlsxwriter.Workbook('fileGenerated/hh.xlsx')
# wr = rootBook.add_worksheet()
# wr.write(0,0,"hello")
# print(type(wr))
# # rootBook.close()
# # sheet = openpyxl.worksheet.worksheet.Worksheet
# worksheet = xlsxwriter.worksheet.Worksheet
# def aroundTo(x: int, num):
#     y = x % num
#     if y != 0:
#         k = x + num - y
#         return k
#     else:
#         return x
#
#
# x = [1, 2, 3, 5, 9, 10, 20, 30]
# for i in x:
#     print(aroundTo(i, 12))
#
# def xlssx(t: x, r: y):
#     t = t.cell(row=5, column=7).value
#     r.write(0, 0, t)
#     print('hello from sheet')
#
#
# xlssx(1,2)
# import operator
#
# az ={'rrrrr':144,'hhhhhh':568,'yyyyyyyyy':144}
# sortedOne =dict(sorted(az.items(), key=operator.itemgetter(1)))
#
# sh=sortedOne.keys()
# print(sortedOne)
# for sh in sh :
#     print(sh)
#                     wr.write(p, column, state, border)
#                     column = column + 1
#                     # CAS VALUE
#                     x = nextBoiteSheet.cell(row=s, column=7).value
#                     wr.write(p, column, x, cassette)
#                     column = column + 1
#                     if state != 'A STOCKER' and state != 'LIBRE':
#                         # TUBE VALUE
#                         x = nextBoiteSheet.cell(row=s, column=10).value
#                         wr.write(p, column, x, stringCassette(str(x)))
#                         column = column + 1
#                         # FIBRE VALUE
#                         x = nextBoiteSheet.cell(row=s, column=9).value
#                         wr.write(p, column, x, stringCassette(str(x)))
#                         column = column + 1
#                         # CABLE VALUE 2
#                         x = nextBoiteSheet.cell(row=s, column=14).value
#                         wr.write(p, column, x, border)
#                         column = column + 1
#                         # BOITE VALUE 2
#                         x = str(nextBoiteSheet.cell(row=s, column=14).value)
#                         boite = x[-4:]
#                         if boite is not None:
#                             boit = getBoiteName(boite)
#                             wr.write(p, column, boit, border)
#                             column = column + 1
#                             newBoite = str(boit)
#                     else:
#                         done = False
from openpyxl import load_workbook

# book =load_workbook('fileGenerated/51074-1.xlsx')
# print(book.sheetnames)
# sheet =book.active
# for row in sheet['E2':'P2']:
#
#     sheet["E2"] = 'F09892110321'
#
# book.save('55.xlsx')
import openpyxl

# for i in range(65, 76):
#     print(chr(i))
# def aroundToThree(x: int):
#     y = x % 3
#     if y != 0:
#         k = x + 3 - y
#         return k
#     else:
#         return x
# print(aroundToThree(0))
# import datetime
#
# date = datetime.datetime.now()
# now = date.strftime("%m/%Y")
# print(now)
#
#
# def aroundTo(x: int, num):
#     y = x % num
#     if y != 0:
#         k = x + num - y
#         return int((k / num) - 1)
#     else:
#         return int((x / num) - 1)
#
#
# diameter = {12: 6, 24: 8.5, 36: 8.5, 48: 9.5, 72: 10.5, 96: 11.5, 144: 11.5, 288: 14.5}
# diameter.update({24:2.5})
# print(diameter[24])
# def is_palindrome(input_string):
#     # We'll create two strings, to compare them
#     new_string = ""
#     reverse_string = ""
#     # Traverse through each letter of the input string
#     for i in input_string:
#         # Add any non-blank letters to the
#         # end of one string, and to the front
#         # of the other string.
#         if i != "":
#             new_string = new_string + i
#             reverse_string = i + reverse_string
#     # Compare the strings
#     if new_string == reverse_string:
#         return True
#     return False
#
#
# print(is_palindrome("abc"))  # Should be False
# print(is_palindrome("kayak"))  # Should be True
# import openpyxl
#
# from os import walk
#
# # the folder source
# monRepertoire = r'C:/Users/etudes20/Desktop/tesstr/FOA SRO 262/'
# wb = openpyxl.load_workbook('C:/Users/etudes20/Desktop/tesstr/85_018_262_POINT_TECHNIQUE_C.xlsx')
# ws = wb.active
# Names = []
# max = ws.max_row
# for n in range(2, max + 1):
#     k = str(ws.cell(n, 1).value)
#     Names.append(k)
# listeFichiers = []
# for (repertoire, sousRepertoires, fichiers) in walk(monRepertoire):
#     listeFichiers.extend(fichiers)
# i = 0
# listName = []
# while i < len(listeFichiers):
#     x = str(listeFichiers[i])
#     x = x[0:-5].strip()
#     if len(x)>11:
#         x = x[0:10].strip()
#     listName.append(x)
#     i += 1
# test = {}
# f = 0
# N = 0
# print(listName)
# for c in Names:
#     found = "NF"
#     for j in listName:
#         if c == j:
#             found = "F"
#     if found.startswith("F"):
#         f += 1
#     else:
#         N += 1
#     print("le point tech " + c + f" is {found} on folder names ")
# print(f"the number of element found is {f} and the number of element not found is {N}")


# nbPrise.append(zaPboDbl.records[k]['nb_prise'])
# tECHNO.append(zaPboDbl.records[k]['techno'])
# typeBat.append(zaPboDbl.records[k]['type_bat'])
# statut.append(zaPboDbl.records[k]['statut'])
#
# print(26 % 12)
# import tkinter as tk
# window = tk.Tk()
# lbl = tk.Label(text="firstStart")
# lbl.pack()