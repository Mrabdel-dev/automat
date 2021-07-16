import editpyxl
from openpyxl import load_workbook

# load the C6 book
workbook = load_workbook('F99999JJMMAA_C6 TEST boite.xlsx')
BookC6 = workbook['Export 2']
# load the C3A book
wb = editpyxl.Workbook()
wb.open('C3A.xlsx')
BookC3A = wb['Commandes Fermes']
###################################################################################
codeInsee = str(BookC6.cell(row=3, column=7).value)
NappuiOrigi = []
typeApp = []
typeEL = {}
NomCab = []
longeur = []
NappuiDist = []
typeAPP = {}
poseAPP = {}

maxRow = BookC6.max_row
x = 0
y = 0
for i in range(9, maxRow):
    N = BookC6.cell(row=i, column=1).value
    T = BookC6.cell(row=i, column=2).value
    if N is not None and T is not None:
        x = N
        y = str(T)
        NappuiOrigi.append(x)
        typeApp.append(y)
        t = str(BookC6.cell(row=i, column=31).value)
        typeAPP.update({x: t})
        p = str(BookC6.cell(row=i, column=36).value)
        poseAPP.update({x: p})
    else:
        NappuiOrigi.append(x)
        typeApp.append(y)
    if y == 'ORT' or y == 'EDF':
        typeEL.update({x: 'AT'})
    else:
        typeEL.update({x: 'A'})
    NomCab.append(str(BookC6.cell(row=i, column=19).value))
    longeur.append(BookC6.cell(row=i, column=20).value)
    NappuiDist.append(BookC6.cell(row=i, column=25).value)

diameter = {12: 6, 24: 8.5, 36: 8.5, 48: 9.5, 72: 10.5, 96: 11.5, 144: 11.5, 288: 14.5}

for item in typeAPP.values():
    print(item)

print('#############', str(BookC6.cell(row=39, column=31).value))


def checkFill(NomDist, index):
    test = True
    for i in range(0, index):
        if NomDist == NappuiOrigi[i]:
            test = False
            break
    return test


def getType(nappui):
    try:
        type = typeEL[nappui]
        return type
    except KeyError:
        if len(str(nappui)) > 5:
            type = 'A'
            return type
        else:
            type = 'AT'
            return type


def getBoitePose(nappui):
    try:
        pose = poseAPP[nappui]
        return pose
    except KeyError:
        pose = None
        return pose


def getTypeAPP(nappui):
    try:
        type = typeAPP[nappui]
        return type
    except KeyError:
        type = None
        return type


def checkPose(boite):
    test = False
    if str(boite) == 'PB' or str(boite) == 'PEO':
        test = True
        return test
    return test


ariene = 'aérien'
msg = "oui remplacement appui"
Len = len(NomCab)
Lin = 15
for i in range(0, Len):
    numC = NomCab[i]
    if numC.startswith('A'):
        TypeA = str(getType(NappuiOrigi[i]))
        TypeB = str(getType(NappuiDist[i]))
        typeAppA = getTypeAPP(NappuiOrigi[i])
        typeAppB = getTypeAPP(NappuiDist[i])
        boitA = getBoitePose(NappuiOrigi[i])
        boitB = getBoitePose(NappuiDist[i])
        diaM = int(numC[9:-4])
        orig = codeInsee + "/" + str(NappuiOrigi[i])
        test = checkFill(NappuiDist[i], i)
        dist = codeInsee + "/" + str(NappuiDist[i])

        if TypeA == 'AT':
            orig = ''
        if TypeB == 'AT':
            dist = ''
        if test:
            if TypeA == 'AT' and TypeB == 'AT':
                pass
            else:
                if checkPose(boitA):
                    if str(boitA) == 'PB':
                        BookC3A.cell(row=Lin, column=14).value = 'A PB Appui'
                        poseAPP.update({NappuiOrigi[i]: None})
                    elif str(boitA) == 'PEO':
                        BookC3A.cell(row=Lin, column=14).value = 'A PEO'
                        poseAPP.update({NappuiOrigi[i]: None})
                elif checkPose(boitB):
                    if str(boitB) == 'PB':
                        BookC3A.cell(row=Lin, column=14).value = 'B PB Appui'
                        poseAPP.update({NappuiDist[i]: None})
                    elif str(boitB) == 'PEO':
                        BookC3A.cell(row=Lin, column=14).value = 'B PEO'
                        poseAPP.update({NappuiDist[i]: None})

                BookC3A.cell(row=Lin, column=2).value = TypeA
                BookC3A.cell(row=Lin, column=3).value = orig
                BookC3A.cell(row=Lin, column=4).value = TypeB
                BookC3A.cell(row=Lin, column=5).value = dist
                BookC3A.cell(row=Lin, column=6).value = longeur[i]
                BookC3A.cell(row=Lin, column=7).value = ariene
                BookC3A.cell(row=Lin, column=10).value = diameter[diaM]
                if typeAppA == 'Remplacement':
                    BookC3A.cell(row=Lin, column=12).value = msg
                    typeAPP.update({NappuiOrigi[i]: None})
                elif typeAppB == 'Remplacement':
                    BookC3A.cell(row=Lin, column=13).value = msg
                    typeAPP.update({NappuiDist[i]: None})

                Lin += 1

wb.save('C3A-NEW.xls')
wb.close()
