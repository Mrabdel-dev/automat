import operator
import string
import openpyxl.worksheet.worksheet
import xlsxwriter.worksheet
import xlsxwriter
from openpyxl import load_workbook
import openpyxl

# generic
pdsSheet = openpyxl.worksheet.worksheet.Worksheet
rootSheet = xlsxwriter.worksheet.Worksheet
# ################define the path of output routage file######################################
rootBook = xlsxwriter.Workbook('fileGenerated/root.xlsx')
wr = rootBook.add_worksheet()
# define the character and style of cell inside excel
bold = rootBook.add_format({'bold': True, "border": 1})
bold1 = rootBook.add_format({'bold': True})
border = rootBook.add_format({"border": 1})
header = rootBook.add_format({'bold': True, 'border': 1, 'bg_color': '#037d50'})
cassette = rootBook.add_format({"bg_color": '#A9A9A9', "border": 1})
cell_formatCapacity = rootBook.add_format({"bg_color": '#E6E6FA', "border": 1})
cell_format1 = rootBook.add_format({"bg_color": 'red', "border": 1})
cell_format2 = rootBook.add_format({"bg_color": 'blue', "border": 1})
cell_format3 = rootBook.add_format({"bg_color": '#00FF00', "border": 1})
cell_format4 = rootBook.add_format({"bg_color": 'yellow', "border": 1})
cell_format5 = rootBook.add_format({"bg_color": '#BF00FF', "border": 1})
cell_format6 = rootBook.add_format({"bg_color": 'white', "border": 1})
cell_format7 = rootBook.add_format({"bg_color": '#FFBF00', "border": 1})
cell_format8 = rootBook.add_format({"bg_color": '#828282', "border": 1})
cell_format9 = rootBook.add_format({"bg_color": '#816B56', "border": 1})
cell_format10 = rootBook.add_format({"bg_color": '#333333', "border": 1})
cell_format11 = rootBook.add_format({"bg_color": '#00FFBF', "border": 1})
cell_format12 = rootBook.add_format({"bg_color": '#FFAAD4', "border": 1})
colorList = [cell_format1, cell_format2, cell_format3, cell_format4, cell_format5, cell_format6, cell_format7,
             cell_format8, cell_format9, cell_format10, cell_format11, cell_format12, border]


def stringCassette(x: string):
    if x.isdigit():
        j = 0
        if int(x) % 12 == 0:
            x = 12
        else:
            x = int(x) % 12

        for i in range(0, 13):
            if i == x:
                x = i
                j = 1
        if j == 1:
            return colorList[x - 1]
        else:
            return colorList[12]
    return colorList[len(colorList) - 1]


def baseHeader():
    wr.write('A1', 'SRO', header)
    wr.write('B1', 'P', header)
    wr.write('C1', 'C ', header)
    wr.write('D1', 'L', header)
    wr.write('E1', 'TIROIR', header)
    wr.write('F1', 'TYPE', header)


def getBoiteName(cableDigit):
    for b in pdsSheets:
        if b.endswith(cableDigit):
            return b


def getSheetName(boite):
    for sh in pdsSheets:
        if sh == boite:
            return sh
    else:
        return print('eroor')


def normalHeader(j):
    for i in range(j, 20):
        wr.write(0, j, 'CAS', header)
        j = j + 1
        wr.write(0, j, 'T', header)
        j = j + 1
        wr.write(0, j, 'F', header)
        j = j + 1
        wr.write(0, j, 'CABLE', header)
        j = j + 1
        wr.write(0, j, 'BOITE', header)
        j = j + 1
        wr.write(0, j, 'TYPE', header)
        j = j + 1


def getTypeI(worksheet: rootSheet, pdsBoite: pdsSheet, lin: int, col: int, pdsline: int):
    x = str(pdsBoite.cell(row=pdsline + 12, column=8).value)
    worksheet.write(lin, col, x)
    col += 1


def getCasI(worksheet: rootSheet, pdsBoite: pdsSheet, lin: int, col: int, pdsline: int):
    x = str(pdsBoite.cell(row=pdsline + 12, column=7).value)
    worksheet.write(lin, col, x, cassette)
    col += 1


def getTubeI(worksheet: rootSheet, pdsBoite: pdsSheet, lin: int, col: int, pdsline: int):
    x = str(pdsBoite.cell(row=pdsline + 12, column=5).value)
    worksheet.write(lin, col, x, stringCassette(x))
    col += 1


def getFibreI(worksheet: rootSheet, pdsBoite: pdsSheet, lin: int, col: int, pdsline: int):
    x = str(pdsBoite.cell(row=pdsline + 12, column=6).value)
    worksheet.write(lin, col, x, stringCassette(x))
    col += 1


def getCableI(worksheet: rootSheet, pdsBoite: pdsSheet, lin: int, col: int, pdsline: int):
    x = str(pdsBoite.cell(row=pdsline + 12, column=1).value)
    worksheet.write(lin, col, x)
    col += 1


def getBoiteI(worksheet: rootSheet, pdsBoite: pdsSheet, lin: int, col: int, pdsline: int):
    x = pdsBoite.cell(row=7, column=1).value
    worksheet.write(lin, col, x, border)
    col += 1


def getTubeII(worksheet: rootSheet, pdsBoite: pdsSheet, lin: int, col: int, pdsline: int):
    x = str(pdsBoite.cell(row=pdsline + 12, column=10).value)
    worksheet.write(lin, col, x, stringCassette(x))
    col += 1


def getFibreII(worksheet: rootSheet, pdsBoite: pdsSheet, lin: int, col: int, pdsline: int):
    x = str(pdsBoite.cell(row=pdsline + 12, column=9).value)
    worksheet.write(lin, col, x, stringCassette(x))
    col += 1


def getCableII(worksheet: rootSheet, pdsBoite: pdsSheet, lin: int, col: int, pdsline: int):
    cable = str(pdsBoite.cell(row=pdsline + 12, column=14).value)
    worksheet.write(lin, col, cable)
    col += 1
    return cable[-4:]


def getBoiteII(worksheet: rootSheet, pdsBoite: pdsSheet, lin: int, col: int, pdsline: int):
    x = pdsBoite.cell(row=7, column=1).value
    worksheet.write(lin, col, x, border)
    col += 1


def getType(worksheet: rootSheet, pdsBoite: pdsSheet, lin: int, col: int, pdsline: int):
    vType = str(pdsBoite.cell(row=pdsline + 12, column=8).value)
    if vType == 'A STOCKER' or vType.endswith('KER') or vType.startswith('A STO'):
        worksheet.write(lin, col, vType)
        pdsline += 1
        col += 1
        return False
    else:
        worksheet.write(lin, col, vType)
        col += 1
        return True


# ##################declare the input PDS file ############################
pdsBook = load_workbook('fileGenerated/PDS.xlsx')
pdsSheets = pdsBook.sheetnames
# boite base of sro
sheetSro = []
# cable base sro
cableSro = []
# capacity of the cable sro
capSro = []
# sro name
SRO = ''
routeDec = {}
srodict = {}
# loop to the boite inside the pds to know SRO cable
for sh in pdsSheets:
    sheet = pdsBook[sh]
    value = sheet.cell(row=1, column=1).value
    if str(value).startswith('SRO'):
        j = 0
        for i in range(12, sheet.max_row + 1):
            ETAT = str(sheet.cell(row=i, column=8).value)
            if ETAT != 'LIBRE' or ETAT != '':
                j = j + 1
        srodict.update({sh: j})
        boiteList = []
        old = 0
        for i in range(12, sheet.max_row):
            cableVal = str(sheet.cell(row=i, column=14).value)
            if cableVal != old and str(cableVal) != '':
                boitBring = cableVal[-4:]
                for s in pdsSheets:
                    boite = str(s)
                    if boite.endswith(boitBring):
                        boiteList.append(s)
                        break
            old = cableVal
        routeDec.update({sh: boiteList})
        SRO = value
        cableSro.append(str(sheet.cell(row=12, column=1).value))
        capSro.append(int(sheet.cell(row=12, column=3).value))

# print(SRO)
# print(routeDec)
# print(srodict)
sortedSro = dict(sorted(srodict.items(), key=operator.itemgetter(1)))
sheetSro = list(sortedSro.keys())
r = sheetSro[0]
sheetSro[0] = sheetSro[1]
sheetSro[1] = r
# print(sheetSro)

# ####################################################### constant declaration if
p = 0
c = []
for i in range(65, 77):
    c.append(chr(i))
c.append(chr(78))
rowWithout = 0  # row without the libre value
l = 0
Line = 1
column = 6
L = 1
T = 0
col = 0
f = 0
pdsLine = 0
Max =0
Len = 0
nextLin = 0
isNext = True
nextBoite = ''
nextCable = ''
libVal = 0
# ##################### the start of copy the values
baseHeader()
normalHeader(6)
for sh in sheetSro:
    sroBoite = pdsBook[sh]
    done = True
    T += 1
    trieur = 'TIROIR_' + str(T)
    for i in range(12, sroBoite.max_row + 1):
        t = str(sroBoite.cell(row=i, column=8).value)
        if t == 'LIBRE' or t == 'PASSAGE':
            libVal = libVal + 1
    Max = sroBoite.max_row - libVal - 11
    while Line < Max:
        if f == 13:
            f = 0
        wr.write('A' + str(Line + 1), SRO, border)
        wr.write('B' + str(Line + 1), Line - 1, border)
        wr.write('E' + str(Line + 1), trieur, border)
        wr.write('F' + str(Line + 1), 'CONNECTEUR', border)
        wr.write('C' + str(Line + 1), c[f], border)
        wr.write('D' + str(L), L, border)
        f = f + 1
        if Line + 1 % 24 == 0:
            if L == 6:
                L = 1
            elif L < 6:
                L = L + 1
        # CAS VALUE
        x = sroBoite.cell(row=Line, column=5).value
        wr.write(Line, column, x, cassette)
        column = column + 1
        # TUBE VALUE
        getTubeI(wr, sroBoite, Line, col, Line)
        # FIBRE VALUE
        getFibreI(wr, sroBoite, Line, col, Line)
        # CABLE VALUE
        getCableI(wr, sroBoite, Line, col, Line)
        # BOITE VALUE
        getBoiteI(wr, sroBoite, Line, col, Line)
        # TYPE VALUE
        getTypeI(wr, sroBoite, Line, col, Line)
        # CAS VALUE
        getCasI(wr, sroBoite, Line, col, Line)
        # TUBE2 VALUE
        getTubeII(wr, sroBoite, Line, col, Line)
        # FIBRE2 VALUE
        getFibreII(wr, sroBoite, Line, col, Line)
        # CABLE2 VALUE 2
        nextCable = getCableII(wr, sroBoite, Line, col, Line)
        # depart to go to the boite next
        nextBoite = getBoiteName(nextCable)
        if nextBoite is not None:
            isNext = True
            pdsLine = Line
            while isNext:
                try:
                    nextSheet = pdsBook[nextBoite]
                    # next sheet Boite name
                    getBoiteII(wr, nextSheet, Line, col, pdsLine)
                    # next type
                    nextType = getType(wr, nextSheet, Line, col, pdsLine)
                    if not nextType:
                        # get next cassete
                        getCasI(wr, nextSheet, Line, col, pdsLine)
                        break
                    else:
                        # get next cassete
                        getCasI(wr, nextSheet, Line, col, pdsLine)
                        # get next tube
                        getTubeII(wr, nextSheet, Line, col, pdsLine)
                        # get next fibre
                        getFibreII(wr, nextSheet, Line, col, pdsLine)
                        # get next cable
                        nextCable = getCableII(wr, nextSheet, Line, col, pdsLine)
                    nextBoite = getBoiteName(nextCable)
                    if nextBoite is not None:
                        continue
                    else:
                        break
                except KeyError:
                    print('boite not found mybe ')
                    isNext = False
        Line += 1
rootBook.close()
