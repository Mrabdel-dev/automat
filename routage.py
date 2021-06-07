import operator
import string

import xlsxwriter
from openpyxl import load_workbook

# ################define the path of output routage file######################################
rootBook = xlsxwriter.Workbook('fileGenerated/root.xlsx')
wr = rootBook.add_worksheet()
# define the character and style of cell inside excel
bold = rootBook.add_format({'bold': True, "border": 1})
bold1 = rootBook.add_format({'bold': True})
border = rootBook.add_format({"border": 1})
header = rootBook.add_format({'bold': True, 'border': 1, 'bg_color': '#037d50'})
cassette = rootBook.add_format({"bg_color": '#A9A9A9', "border": 1})
cell_formatCapacity = rootBook.add_format({"bg_color": '#E6E6FA', "border": 1})
cell_format1 = rootBook.add_format({"bg_color": 'red', "border": 1})
cell_format2 = rootBook.add_format({"bg_color": 'blue', "border": 1})
cell_format3 = rootBook.add_format({"bg_color": '#00FF00', "border": 1})
cell_format4 = rootBook.add_format({"bg_color": 'yellow', "border": 1})
cell_format5 = rootBook.add_format({"bg_color": '#BF00FF', "border": 1})
cell_format6 = rootBook.add_format({"bg_color": 'white', "border": 1})
cell_format7 = rootBook.add_format({"bg_color": '#FFBF00', "border": 1})
cell_format8 = rootBook.add_format({"bg_color": '#828282', "border": 1})
cell_format9 = rootBook.add_format({"bg_color": '#816B56', "border": 1})
cell_format10 = rootBook.add_format({"bg_color": '#333333', "border": 1})
cell_format11 = rootBook.add_format({"bg_color": '#00FFBF', "border": 1})
cell_format12 = rootBook.add_format({"bg_color": '#FFAAD4', "border": 1})
colorList = [cell_format1, cell_format2, cell_format3, cell_format4, cell_format5, cell_format6, cell_format7,
             cell_format8, cell_format9, cell_format10, cell_format11, cell_format12, border]


def stringCassette(x: string):
    if x.isdigit():
        j = 0
        if int(x) % 12 == 0:
            x = 12
        else:
            x = int(x) % 12

        for i in range(0, 13):
            if i == x:
                x = i
                j = 1
        if j == 1:
            return colorList[x - 1]
        else:
            return colorList[12]
    return colorList[len(colorList) - 1]


def baseHeader():
    wr.write('A1', 'SRO', header)
    wr.write('B1', 'P', header)
    wr.write('C1', 'C ', header)
    wr.write('D1', 'L', header)
    wr.write('E1', 'TIROIR', header)
    wr.write('F1', 'TYPE', header)


def getBoiteName(cabledigite):
    for b in pdsSheets:
        if b.endswith(cabledigite):
            return b


def getSheetName(boite):
    for sh in pdsSheets:
        if sh == boite:
            return sh
    else:
        return print('eroor')


def normalHeader(j):
    for i in range(j, 20):
        wr.write(0, j, 'CAS', header)
        j = j + 1
        wr.write(0, j, 'T', header)
        j = j + 1
        wr.write(0, j, 'F', header)
        j = j + 1
        wr.write(0, j, 'CABLE', header)
        j = j + 1
        wr.write(0, j, 'BOITE', header)
        j = j + 1
        wr.write(0, j, 'TYPE', header)
        j = j + 1


trysh = []
# ##################declare the input PDS file ############################
pdsBook = load_workbook('fileGenerated/PDS.xlsx')
pdsSheets = pdsBook.sheetnames
# boite base of sro
sheetSro = []
# cable base sro
cableSro = []
# capacity of the cable sro
capSro = []
# sro name
SRO = ''
routeDec = {}
srodict = {}
# loop to the boite inside the pds to know SRO cable
for sh in pdsSheets:
    sheet = pdsBook[sh]
    value = sheet.cell(row=1, column=1).value
    if str(value).startswith('SRO'):
        j = 0
        for i in range(12, sheet.max_row + 1):
            ETAT = str(sheet.cell(row=i, column=8).value)
            if ETAT != 'LIBRE' or ETAT != '':
                j = j + 1
        srodict.update({sh: j})
        boiteList = []
        old = 0
        for i in range(12, sheet.max_row):
            cableVal = str(sheet.cell(row=i, column=14).value)
            if cableVal != old and str(cableVal) != '':
                boitBring = cableVal[-4:]
                for s in pdsSheets:
                    boit = str(s)
                    if boit.endswith(boitBring):
                        boiteList.append(s)
                        break
            old = cableVal
        routeDec.update({sh: boiteList})
        SRO = value
        cableSro.append(str(sheet.cell(row=12, column=1).value))
        capSro.append(int(sheet.cell(row=12, column=3).value))

print(SRO)
print(routeDec)
print(srodict)
sortedSro = dict(sorted(srodict.items(), key=operator.itemgetter(1)))
sheetSro = sortedSro.keys()
print(sheetSro)
baseHeader()
normalHeader(6)
p = 0
c = []
for i in range(65, 77):
    c.append(chr(i))
c.append(chr(78))
L = 0
T = 0
f = 0
N = 0
Len = 0
for b in sheetSro:
    bshet = pdsBook[b]
    L = 1
    T = T + 1
    v = 0
    for i in range(12, bshet.max_row + 1):
        t = str(bshet.cell(row=i, column=8).value)
        if t == 'LIBRE' or t == 'PASSAGE':
            v = v + 1
    print('#' * 25)
    print(b)
    print(bshet.max_row - 11)
    print(v)
    Len = Len + bshet.max_row - v - 11
    col = 6
    # define the base element
    for p in range(N + 2, Len + 2):
        if f == 13:
            f = 0
        wr.write('A' + str(p), SRO, border)
        wr.write('B' + str(p), p - 1, border)
        wr.write('E' + str(p), 'TIROIR_' + str(T), border)
        wr.write('F' + str(p), 'CONNECTEUR', border)
        wr.write('C' + str(p), c[f], border)
        wr.write('D' + str(p), L, border)
        f = f + 1
        if p % 24 == 0:
            if L == 6:
                L = 1
            elif L < 6:
                L = L + 1
    shRow = []
    for rw in range(12, bshet.max_row + 1):
        raw = bshet.cell(row=rw, column=8).value
        if str(raw) == 'LIBRE' or str(raw) == 'PASSAGE':
            continue
        else:
            shRow.append(rw)

    # full up the table with value
    for p, s in zip(range(N + 1, Len + 2), shRow):
        column = col
        # CAS VALUE
        x = bshet.cell(row=s, column=5).value
        wr.write(p, column, x, cassette)
        column = column + 1
        # TUBE VALUE
        x = bshet.cell(row=s, column=5).value
        wr.write(p, column, x, stringCassette(str(x)))
        column = column + 1
        # FIBRE VALUE
        x = bshet.cell(row=s, column=6).value
        wr.write(p, column, x, stringCassette(str(x)))
        column = column + 1
        # CABLE VALUE
        x = bshet.cell(row=12, column=1).value
        wr.write(p, column, x, border)
        column = column + 1
        # BOITE VALUE
        x = bshet.cell(row=7, column=1).value
        wr.write(p, column, x, border)
        column = column + 1
        # TYPE VALUE
        x = bshet.cell(row=s, column=8).value
        wr.write(p, column, x, border)
        column = column + 1
        # CAS VALUE
        x = bshet.cell(row=s, column=7).value
        wr.write(p, column, x, cassette)
        column = column + 1
        # TUBE VALUE
        x = bshet.cell(row=s, column=10).value
        wr.write(p, column, x, stringCassette(str(x)))
        column = column + 1
        # FIBRE VALUE
        x = bshet.cell(row=s, column=9).value
        wr.write(p, column, x, stringCassette(str(x)))
        column = column + 1
        # CABLE VALUE 2
        x = bshet.cell(row=s, column=14).value
        wr.write(p, column, x, border)
        column = column + 1
        # BOITE VALUE 2
        x = str(bshet.cell(row=s, column=14).value)
        boite = x[-4:]
        if boite is not None:
            boit = getBoiteName(boite)
            wr.write(p, column, boit, border)
            column = column + 1
            done = True
            newBoite = str(boit)
            trysh.append(newBoite)
        if newBoite is not None:
            while done:
                try:
                    nextBoiteSheet = pdsBook[newBoite]
                    # TYPE VALUE
                    state = nextBoiteSheet.cell(row=s, column=8).value
                    wr.write(p, column, state, border)
                    column = column + 1
                    # CAS VALUE
                    x = nextBoiteSheet.cell(row=s, column=7).value
                    wr.write(p, column, x, cassette)
                    column = column + 1
                    if state != 'A STOCKER' and state != 'LIBRE':
                        # TUBE VALUE
                        x = nextBoiteSheet.cell(row=s, column=10).value
                        wr.write(p, column, x, stringCassette(str(x)))
                        column = column + 1
                        # FIBRE VALUE
                        x = nextBoiteSheet.cell(row=s, column=9).value
                        wr.write(p, column, x, stringCassette(str(x)))
                        column = column + 1
                        # CABLE VALUE 2
                        x = nextBoiteSheet.cell(row=s, column=14).value
                        wr.write(p, column, x, border)
                        column = column + 1
                        # BOITE VALUE 2
                        x = str(nextBoiteSheet.cell(row=s, column=14).value)
                        boite = x[-4:]
                        if boite is not None:
                            boit = getBoiteName(boite)
                            wr.write(p, column, boit, border)
                            column = column + 1
                            newBoite = str(boit)
                    else:
                        done = False
                except KeyError:
                    print('one empthy value passed')
                    done=False

    N = Len
for i in trysh:
    try:
        sheet = pdsBook[i]
    except KeyError:
        print(i)
rootBook.close()
