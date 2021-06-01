import xlsxwriter
from openpyxl import load_workbook
#looad your pds file here
pds =load_workbook('PDS.xlsx')
wpds= pds.sheetnames

file="epes.xlsx"
eps =load_workbook(file)
epsheet = eps.active
k = epsheet.max_row
print(k)
################## the part of coping values from pds to new file ######################
for s in wpds:
    sheet = pds[s]
    MaxRow= sheet.max_row
    MaxCol= sheet.max_column
    k = epsheet.max_row
    print('#'*15)
    print(k)
    for i in range(12,MaxRow+1):
        k = k + 1
        for j in range(1,MaxCol+1):

            #read valueus
            valin = sheet.cell(row=i,column=j).value
            #append the value
            epsheet.cell(row=k, column=j).value=valin
file2='uapdte.xlsx'
eps.save(file2)
