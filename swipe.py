from dbfread import DBF
from openpyxl import load_workbook
import datetime

# date configuration
now = datetime.datetime.now()
date = now.strftime("%m/%Y")
# load your pds file here
fciTable = DBF('fCI-011.dbf', load=True, encoding='iso-8859-1')
pointTechTable = DBF('21_011_POINT_TECHNIQUE_A.dbf', load=True, encoding='iso-8859-1')
# FROM THE TECHNIC POINT
pointlen = len(pointTechTable)
pointNom = []
pointCode = []
pointFonc = []
pointStruc = []
pointPrp = []
for p in range(0, pointlen):
    pointNom.append(pointTechTable.records[p]['NOM'])
    pointCode.append(pointTechTable.records[p]['CODE'])
    pointFonc.append(pointTechTable.records[p]['TYPE_FONC'])
    pointStruc.append(pointTechTable.records[p]['RATTACH'])
    pointPrp.append(pointTechTable.records[p]['PROPRIETAI'])
# FROM THE FCI 
fciNom = []
fciCode = []
fcilen = len(fciTable)
for f in range(0, fcilen):
    try:
        fciNom.append(fciTable.records[f]['POTEAU_CHA'])
    except KeyError:
        fciNom.append(fciTable.records[f]['N__'])

    fciCode.append(fciTable.records[f]['FCI'])
etiquete = load_workbook('NRO-21_011_DETAIL_ETIQUETTE.xlsx')
listsh = etiquete.sheetnames


def getPointtype(point):
    try:
        index = pointNom.index(point)
        prop = pointPrp[index]
        return prop
    except:
        prop = 'point not found'
        return prop


def getSroname(point):
    try:
        index = pointNom.index(point)
        sro = str(pointStruc[index])
        return sro
    except:
        return 'point not found'


for sh in listsh:
    sheet = etiquete[sh]
    maxrow = sheet.max_row
    for i in range(1, maxrow + 1):
        val = str(sheet.cell(row=i, column=1).value)
        typ = str(getPointtype(val))
        sro = getSroname(val)
        try:

            index = fciNom.index(val)
            code = str(fciCode[index])
            sheet.cell(row=i, column=6).value = code + "-" + str(date)
        except:
            if typ.startswith("OR"):
                sheet.cell(row=i, column=6).value = "None"
            elif val.startswith("SR") or val.startswith("NRO"):
                sheet.cell(row=i, column=6).value = val + "-" + str(date)
            else:
                sheet.cell(row=i, column=6).value = sro + "-" + str(date)


etiquete.save('NRO-21_011_DETAIL_ETIQUETTE-NEW.xlsx')