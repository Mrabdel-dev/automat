# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""
from os import walk
from dbfread import DBF
from openpyxl import load_workbook

workbook = load_workbook(r'C:\Users\etudes20\Desktop/SIO-REC-85.xlsx')
sheet = workbook.active
monRepertoire = r'C:\Users\etudes20\Desktop\Nouveau dossier/'

listeFichiers = []
for (repertoire, sousRepertoires, fichiers) in walk(monRepertoire):
    listeFichiers.extend(fichiers)

r = 5
i = 0
while i < len(listeFichiers):
    cableTable = DBF(monRepertoire + listeFichiers[i], load=True, encoding='iso-8859-1')
    cableLen = len(cableTable)
    cableType = []
    cableEate = []
    cableLong = []
    airexi = 0
    airmco = 0
    soutexi = 0
    soutmco = 0
    for j in range(0, cableLen):
        cableType.append(cableTable.records[j]['TYPE_STRUC'])
        cableEate.append(cableTable.records[j]['ETAT'])
        cableLong.append(cableTable.records[j]['LGR_REELLE'])
    for t, e, l in zip(cableType, cableEate, cableLong):
        if str(t).startswith("A"):
            try:
                if str(e).startswith("M"):
                    airmco += float(l)
                elif str(e).startswith("E"):
                    airexi += float(l)
            except:
                print(listeFichiers[i])
        elif str(t).startswith("C"):
            try:
                if str(e).startswith("M"):
                    soutmco += float(l)
                elif str(e).startswith("E"):
                    soutexi += float(l)
            except:
                print(listeFichiers[i])

    sheet.cell(row=r, column=4).value = listeFichiers[i][0:10]
    sheet.cell(row=r, column=5).value = airexi
    sheet.cell(row=r, column=6).value = airmco
    sheet.cell(row=r, column=7).value = soutexi
    sheet.cell(row=r, column=8).value = soutmco

    i += 1
    r += 1
workbook.save(r'C:\Users\etudes20\Desktop/SIO-REC-85.xlsx')
