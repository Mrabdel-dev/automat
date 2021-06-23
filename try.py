# def duplicates(lst, item):
#     return [i for i, x in enumerate(lst) if x == item]
#
#
#
# listi = ['1','2','158','99','93']
# for i in enumerate(listi):
#     print(i)
# c = []
# for i in range(65, 77):
#     c.append(chr(i))
# c.append(chr(78))
# import openpyxl.worksheet.worksheet
# import xlsxwriter.worksheet
from openpyxl import load_workbook


# pdsBook = load_workbook('fileGenerated/PDS.xlsx')
# pdsSheets = pdsBook.sheetnames
# y = pdsBook[pdsSheets[0]]
# print(type(y))
# def sheets(x:openpyxl.worksheet.worksheet.Worksheet):
#     x.cell()
# print(c[15])
# import xlsxwriter
# rootBook = xlsxwriter.Workbook('fileGenerated/hh.xlsx')
# wr = rootBook.add_worksheet()
# wr.write(0,0,"hello")
# print(type(wr))
# # rootBook.close()
# # sheet = openpyxl.worksheet.worksheet.Worksheet
# worksheet = xlsxwriter.worksheet.Worksheet
def aroundTo(x: int, num):
    y = x % num
    if y != 0:
        k = x + num - y
        return k
    else:
        return x


x = [1, 2, 3, 5, 9, 10, 20, 30]
for i in x:
    print(aroundTo(i, 12))
#
# def xlssx(t: x, r: y):
#     t = t.cell(row=5, column=7).value
#     r.write(0, 0, t)
#     print('hello from sheet')
#
#
# xlssx(1,2)
# import operator
#
# az ={'rrrrr':144,'hhhhhh':568,'yyyyyyyyy':144}
# sortedOne =dict(sorted(az.items(), key=operator.itemgetter(1)))
#
# sh=sortedOne.keys()
# print(sortedOne)
# for sh in sh :
#     print(sh)
#                     wr.write(p, column, state, border)
#                     column = column + 1
#                     # CAS VALUE
#                     x = nextBoiteSheet.cell(row=s, column=7).value
#                     wr.write(p, column, x, cassette)
#                     column = column + 1
#                     if state != 'A STOCKER' and state != 'LIBRE':
#                         # TUBE VALUE
#                         x = nextBoiteSheet.cell(row=s, column=10).value
#                         wr.write(p, column, x, stringCassette(str(x)))
#                         column = column + 1
#                         # FIBRE VALUE
#                         x = nextBoiteSheet.cell(row=s, column=9).value
#                         wr.write(p, column, x, stringCassette(str(x)))
#                         column = column + 1
#                         # CABLE VALUE 2
#                         x = nextBoiteSheet.cell(row=s, column=14).value
#                         wr.write(p, column, x, border)
#                         column = column + 1
#                         # BOITE VALUE 2
#                         x = str(nextBoiteSheet.cell(row=s, column=14).value)
#                         boite = x[-4:]
#                         if boite is not None:
#                             boit = getBoiteName(boite)
#                             wr.write(p, column, boit, border)
#                             column = column + 1
#                             newBoite = str(boit)
#                     else:
#                         done = False
from openpyxl import load_workbook

# book =load_workbook('fileGenerated/51074-1.xlsx')
# print(book.sheetnames)
# sheet =book.active
# for row in sheet['E2':'P2']:
#
#     sheet["E2"] = 'F09892110321'
#
# book.save('55.xlsx')
import openpyxl

# for i in range(65, 76):
#     print(chr(i))
# def aroundToThree(x: int):
#     y = x % 3
#     if y != 0:
#         k = x + 3 - y
#         return k
#     else:
#         return x
# print(aroundToThree(0))
