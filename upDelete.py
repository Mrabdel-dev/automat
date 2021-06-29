from openpyxl import load_workbook

# load the old book you gona modify
workbookOldFree = load_workbook('C:/Users/etudes20/Desktop/free/Suivi Projet FREE-IDF 21-06.xlsx')
oldBookFree = workbookOldFree.active
# load the old book you gona modify
workbookOldPaca = load_workbook('C:/Users/etudes20/Desktop/free/Suivi projets Free  PACA 28-06.xlsx')
oldBookPaca = workbookOldPaca.active
# load the new book that you want get value from it
workbookNew = load_workbook('C:/Users/etudes20/Desktop/free/export_projets_2021-06-29-12-13-13.xlsx')
newBook = workbookNew.active
# define parameter for loop
maxRow = newBook.max_row
print('######## the number of values in the new file########')
print(maxRow - 1)
maxCol = newBook.max_column
maxRowOldFree = oldBookFree.max_row
maxRowOldPaca = oldBookPaca.max_row

maxColOldFree = oldBookFree.max_column
maxColOldPaca = oldBookPaca.max_column
NbrAddFree = 0
NbrExsFree = 0
NbrDelFree = 0
NbrAddPaca = 0
NbrExsPaca = 0
NbrDelPaca = 0
# the first loop is add all the new value that doesn't exist in the old file
for i in range(2, maxRow + 1):
    Test = str(newBook.cell(row=i, column=7).value)
    if Test == "Sebastien GELSI":
        valNew = newBook.cell(row=i, column=4).value
        for j in range(2, maxRowOldFree + 1):
            valOld = oldBookFree.cell(row=j, column=4).value
            if valNew == valOld:
                found = 1
                NbrExsFree = NbrExsFree + 1
                break
            else:
                found = 0
        if found == 0:
            modRowOld = oldBookFree.max_row
            u = modRowOld + 1
            NbrAddFree = NbrAddFree + 1
            for k in range(1, 10):
                valN = newBook.cell(row=i, column=k).value
                oldBookFree.cell(row=u, column=k).value = valN
    else:
        valNew = newBook.cell(row=i, column=4).value
        for j in range(2, maxRowOldPaca + 1):
            valOld = oldBookPaca.cell(row=j, column=4).value
            if valNew == valOld:
                found = 1
                NbrExsPaca = NbrExsPaca + 1
                break
            else:
                found = 0
        if found == 0:
            modRowOld = oldBookPaca.max_row
            u = modRowOld + 1
            NbrAddPaca = NbrAddPaca + 1
            for k in range(1, 10):
                valN = newBook.cell(row=i, column=k).value
                oldBookPaca.cell(row=u, column=k).value = valN

newMax = oldBookFree.max_row
# the second loop is for delete all value doesn't come with new file
for c in range(2, newMax):
    Test = str(oldBookFree.cell(row=c, column=6).value)
    if Test == "Sebastien GELSI":
        valTest = oldBookFree.cell(row=c, column=4).value
        for d in range(2, maxRow + 1):
            valTestedWith = newBook.cell(row=d, column=4).value
            if valTest == valTestedWith:
                y = 1
                break
            else:
                y = 0
        if y == 0:
            NbrDelFree = NbrDelFree + 1
            for e in range(1, maxColOldFree + 1):
                oldBookFree.cell(row=c, column=e).value = ''

newMaxPaca = oldBookPaca.max_row
for c in range(2,newMaxPaca):
    Test = str(oldBookPaca.cell(row=c, column=6).value)
    if Test == "Philippe PHILIS":
        print(c)
        valTest = oldBookPaca.cell(row=c, column=4).value
        for d in range(2, maxRow + 1):
            valTestedWith = newBook.cell(row=d, column=4).value
            if valTest == valTestedWith:
                y = 1
                break
            else:
                y = 0
        if y == 0:
            NbrDelPaca = NbrDelPaca + 1
            for e in range(1, maxColOldPaca + 1):
                oldBookPaca.cell(row=c, column=e).value = ''
    else:
        print("yes")
        for e in range(1, maxColOldPaca + 1):
            oldBookPaca.cell(row=c, column=e).value = ''


print('######## the number of values in the old file free ########')
print(maxRowOldFree - 1)
print('##### statistique of operationFree ##########')
print(f'number of element added is : ==>{NbrAddFree}')
print(f'number of element deleted is : ==>{NbrDelFree}')
print(f'number of element exist already is : ==>{NbrExsFree}')
print('#################### the number of values in the old file after update ####################')
print(oldBookFree.max_row - 1)
print('#'*26)
print('######## the number of values in the old file paca ########')
print(maxRowOldPaca - 1)
print('##### statistique of operationPaca ##########')
print(f'number of element added is : ==>{NbrAddPaca}')
print(f'number of element deleted is : ==>{NbrDelPaca}')
print(f'number of element exist already is : ==>{NbrExsPaca}')
print('#################### the number of values in the old file after update ####################')
print(oldBookPaca.max_row - 1)
# save the file
fileNFree = 'fileGenerated/Suivi Projet FREE-IDF-Nouveau.xlsx'
fileNPaca = 'fileGenerated/Suivi projets Free PACA-Nouveau.xlsx'
workbookOldFree.save(fileNFree)
workbookOldPaca.save(fileNPaca)
